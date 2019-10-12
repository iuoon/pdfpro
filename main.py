import os

import openpyxl as openpyxl
import pdfkit
import win32api,win32print
from PIL import ImageFont
from PyPDF2 import PdfFileReader, PdfFileWriter, PdfFileMerger
from jinja2 import Environment, FileSystemLoader
import wx


APP_TITLE = u'标签生成器'
APP_ICON = 'res/python.ico'

class mainFrame(wx.Frame):
    '''程序主窗口类，继承自wx.Frame'''
    def __init__(self, parent):
        '''构造函数'''
        wx.Frame.__init__(self, parent, -1, APP_TITLE)
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.SetSize((600, 480))
        self.Center()
        wx.StaticText(self, -1, u'设置纸张尺寸： 宽', pos=(10, 20), size=(100, -1), style=wx.ALIGN_RIGHT)
        self.pageWidth = wx.TextCtrl(self, -1, '', pos=(110, 18), size=(30, -1), name='pageWidth', style=wx.TE_LEFT)
        self.pageHeight = wx.TextCtrl(self, -1, '', pos=(160, 18), size=(30, -1), name='pageHeight', style=wx.TE_LEFT)
        # 生成html模板
    def generate_html(self,body):
        env = Environment(loader=FileSystemLoader('./'))
        template = env.get_template('temp.html')
        with open("template.html", 'w+', encoding='utf-8') as fout:
            html_content = template.render(body=body)
            fout.write(html_content)
    
    # 生成pdf
    def generate_pdf(self,content):
        # 生成pdf页，并设置大小
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        path_wk = BASE_DIR + '\\w\\wkhtmltopdf.exe'
        config = pdfkit.configuration(wkhtmltopdf=path_wk)
        options = {
            'encoding': "utf-8",
            'page-height': '50mm',  # 设置高
            'page-width': '80mm',  # 设置宽
            'margin-top': '0mm',
            'margin-right': '0mm',
            'margin-bottom': '0mm',
            'margin-left': '0mm'
        }
        pdfkit.from_string(content, "temp.pdf", options=options, configuration=config)
    
        #创建一个pdf文件合并对象
        # pdfMerger = PdfFileMerger()
        # if (os.path.exists(BASE_DIR + "\\result.pdf")):
        #     with open("result.pdf", 'rb') as fe1:
        #         pdfMerger.append(fe1)
        # with open("temp.pdf", 'rb') as fe2:
        #     pdfMerger.append(fe2)
        # with open("result.pdf", 'ab+') as fe3:
        #     pdfMerger.write(fe3)
    
        pdfFileReader = PdfFileReader(open("temp.pdf", 'rb'))
        pdfFileWriter = PdfFileWriter()
    
        if (os.path.exists(BASE_DIR + "\\result.pdf")):
            pdfFileReader0 = PdfFileReader(open("result.pdf", 'rb'))
            for page in range(pdfFileReader0.getNumPages()):
                # 将每页添加到writer对象
                pdfFileWriter.addPage(pdfFileReader0.getPage(page))
        pageObj = pdfFileReader.getPage(0)
        pdfFileWriter.addPage(pageObj)
        pdfFileWriter.write(open("result.pdf", 'ab+'))
    
    
    def start(self):
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        if (os.path.exists(BASE_DIR + "\\result.pdf")):
            os.remove(BASE_DIR + "\\result.pdf")
    
        workbook = openpyxl.load_workbook("data.xlsx")
        sheet = workbook.get_active_sheet()
        title_data = []
        for i in range(1, sheet.max_column + 1):
            title_data.append(sheet.cell(1, i).value)
        self.generate_html(title_data)
    
        for i in range(2, sheet.max_row + 1):
            content = ''
            with open("template.html", 'r', encoding='utf-8') as f:
                content = f.read()
            for j in range(1, sheet.max_column + 1):
                content = content.replace("#"+str(j)+"#", str(sheet.cell(i, j).value))
            self.generate_pdf(content)
    
        # font = ImageFont.truetype(font='simsun.ttc', size=20)
        # text = "hello"
        # width, height = font.getsize(text)
        # print(width, height)
        printers = win32print.GetDefaultPrinter()
        print(printers)

class mainApp(wx.App):
    def OnInit(self):
        self.SetAppName(APP_TITLE)
        self.Frame = mainFrame(None)
        self.Frame.Show()
        return True

if __name__ == "__main__":
    app = mainApp()
    app.MainLoop()